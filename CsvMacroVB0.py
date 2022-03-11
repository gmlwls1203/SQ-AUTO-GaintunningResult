# -- coding:utf-8 --
import statistics
import csv
import datetime
import os
import re
import time
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import PatternFill

import warnings

warnings.filterwarnings('ignore')

MAKE_A_PLOT_FLAG = True

gap_validity = True
gap_validity_text = ""
t2a_validity = True
t2a_validity_text = ""
rng_validity = True
rng_validity_text = ""
acc_validity = True
acc_validity_text = ""
score_result = 0


def gap_set_validity(flag, text):
    global gap_validity
    global gap_validity_text
    gap_validity = flag
    gap_validity_text = text


def t2a_set_validity(flag, text):
    global t2a_validity
    global t2a_validity_text
    t2a_validity = flag
    t2a_validity_text = text


def rng_set_validity(flag, text):
    global rng_validity
    global rng_validity_text
    rng_validity = flag
    rng_validity_text = text


def acc_set_validity(flag, text):
    global acc_validity
    global acc_validity_text
    acc_validity = flag
    acc_validity_text = text


def tar2act_time_spec(spec, diff_time):
    global score_result
    if 0 > diff_time:
        color_matplotlib = 'black'  # move over
        color_openpyxl = 'FF000000'
        score_result += 100
        t2a_set_validity(False, "tar2act move over")

    elif (spec * 1.00) > diff_time:  # 78 * 1 < 229
        color_matplotlib = 'forestgreen'  # normal
        color_openpyxl = 'FF44FF33'
        # SetValidty(True, "")

    elif (spec * 1.05) > diff_time:  # 78 * 1.05 < 229
        color_matplotlib = 'gold'  # caution
        color_openpyxl = 'FFAA9148'
        score_result += 1
        t2a_set_validity(False, "Spec x 1.00 times under")

    elif (spec * 1.10) > diff_time:  # 78 * 1.1 < 229
        color_matplotlib = 'orange'  # warning
        color_openpyxl = 'FFFF9911'
        score_result += 3
        t2a_set_validity(False, "Spec x 1.05 times over")

    else:  # 300 * 1.15 = 345 < input
        color_matplotlib = 'red'  # danger
        color_openpyxl = 'FFFF2211'
        score_result += 5
        t2a_set_validity(False, "Spec x 1.10 times over")

    return [color_matplotlib, color_openpyxl]


def tar2act_time_check(src, dst, diff_time):
    spec = 0.00

    """
    # P-----------------R--------N---------D
    if (src == 'P' and dst == 'D') or (src == 'D' and dst == 'P'):
        spec = 300.00  # P2D, D2P : 300.00ms (angle diff = 37.43)
    elif (src == 'P' and dst == 'N') or (src == 'N' and dst == 'P'):
        spec = 221.21  # P2N, N2P : 221.21ms (angle diff = 27.6)
    elif (src == 'P' and dst == 'R') or (src == 'R' and dst == 'P'):
        spec = 142.42  # R2P, P2R : 142.42ms (angle diff = 17.77)
    elif (src == 'R' and dst == 'D') or (src == 'D' and dst == 'R'):
        spec = 157.57  # D2R, R2D : 157.57ms (angle diff = 19.66)
    elif (src == 'R' and dst == 'N') or (src == 'N' and dst == 'R'):
        spec = 78.79  # R2N, N2R : 78.79ms (angle diff = 9.83)
    elif (src == 'N' and dst == 'D') or (src == 'D' and dst == 'N'):
        spec = 78.79  # N2D, D2N : 78.79ms (angle diff = 9.83)
    else:
        spec = 0
    """

    if src == 'P' and dst == 'D':
        spec = 500.00  # P2D : 300.00ms (angle diff = 37.43)
    elif src == 'P' and dst == 'N':
        spec = 500.00  # P2N : 221.21ms (angle diff = 27.6)
    elif src == 'P' and dst == 'R':
        spec = 500.00  # P2R : 142.42ms (angle diff = 17.77)
    elif src == 'R' and dst == 'P':
        spec = 500.00  # R2P : 142.42ms (angle diff = 17.77)
    elif src == 'R' and dst == 'D':
        spec = 500.00  # R2D : 157.57ms (angle diff = 19.66)
    elif src == 'R' and dst == 'N':
        spec = 500.00  # R2N : 78.79ms (angle diff = 9.83)
    elif src == 'N' and dst == 'P':
        spec = 500.00  # N2P : 221.21ms (angle diff = 27.6)
    elif src == 'N' and dst == 'R':
        spec = 500.00  # N2R : 78.79ms (angle diff = 9.83)
    elif src == 'N' and dst == 'D':
        spec = 500.00  # N2D : 78.79ms (angle diff = 9.83)
    elif src == 'D' and dst == 'P':
        spec = 500.00  # D2P : 300.00ms (angle diff = 37.43)
    elif src == 'D' and dst == 'R':
        spec = 500.00  # D2R : 157.57ms (angle diff = 19.66)
    elif src == 'D' and dst == 'N':
        spec = 500.00  # D2N : 78.79ms (angle diff = 9.83)
    else:
        spec = 0

    color = tar2act_time_spec(float(spec), float(diff_time))

    return [spec, color[0], color[1]]


"""
VA1
1. raw 데이터 규칙 파악 및 파일
2. 폴더 접근 유효성 확인 및 폴더 접근체크


VA1 -> VA2
1. 그래픽 디스플레이 수정(요구사항에 부합하는지 확인 필요)
  - 포지션
  - Target
  - Act
  - 전류
  - 전압
2. 이강석 책임 요구사항 검토


VA2 -> VA3
1. 그래픽 디스플레이 수정
2. 황철민 사원 요구사항 검토
3. 김희진 사원 요구사항 검토
  - UVW 전류 체크
  - 포지션 raw 데이터에서 0점 조절할 수 있는 heejin 변수 추가 (현재는 삭제 처리)


VA3 -> VA4
1. Target, Act 동일 그래프에 출력 및 diff time 체크
2. 엑셀파일 정렬


VA4 -> VA5
1. 배포파일 생성 (실패)
2. Python 설치방법 공유
3. position 데이터 Y-axis 고정 (P : 10도 ~ D : 47도)
4. Raw 데이터 컬러 셋 추가
5. 그래프에 주석 및 시간, 크기 설정


VA5 -> VA6
1. 데이터에 오류가 있는 로깅파일이 있음. (버리긴 아까움)
  - RANSAC 적용하여 초기 위치, 끝 위치 정확히 판단이 필요 함
    : 오류 데이터는 초과전압으로 판단하여 해당 부분은 이전값 유지
2. 프로세싱 성능 확인
  - 입력 파일 및 폴더 validity check (0.0%)
  - 이미지 생성 (94.1%)
  - 이미지 저장 (0.5%)
  - 엑셀 저장 (0.4%)
  - 로데이터 프로세싱 (5.0%)
3. 프로세싱 데이터 최적화
  - 프로세싱 처리속도 개선 - raw 데이터당 1.2s 로그파일 -> 0.8s만 체크 (33% 프로세싱 성능 개선)


VA6 -> VA7 변경 검토 ('20.06/30~07/05)
1. 코드 리팩토링
  - validity 기능 확장 (완)
  - 누구든 유지보수 가능하도록 주석 추가 및 코드 정리 (완)
  - 셀프 코드리뷰 (완)
2. 유저 UI 개선
  - 콘솔에 LV3 이하 삭제 진행 상황 표시 (취소)
  - 엑셀에 문제가 있는 부분은 Color set(완)
    : 빠른 분석을 위함
  - 이미지 분류 개선(Good / Bad :: PR, PD, PN, ...) (완)


VA7 -> VA8 변경 검토 ('20.07/21~07/28)
1. 알고리즘 보강
  - Ovs Gap error 추가 (완)
  - 언더슈트 체크 추가 (완)
    : 오버슈트 포지션과 셋 포지션사이 외 데이터가 10ms 정도 이상 연속으로 들어오면 언더슈트로 체크, Validity 추가
  - 800ms 기준 포지션 정확도 체크(완)
    : 열 추가 + 1
2. 전류 흐르는 시간을 측정하는 알고리즘 추가 (완)
  - 전자파 특성 고려 - 제어시간 측정을 어떻게 하지..?
    : 전류가 안정화 되지 안았을 때를 찾으면 되는데, 전류의 크기 변화는 작다.
    : 전류와 전압파형이 동일한 위상에서 변화가 있으며, 전압은 변화가 크다.
    : 전압의 시작부터 raw 데이터 100개 샘플 평균값을 기준으로
    : 시작 포인트부터 0.3V, 끝 포인트부터 0.3V 차이를 탐색하여 컨트롤 시간 측정
    : 열 추가 + 1
3. 엑셀 시트 정렬 및 그룹핑(완)
  - GAIN 별 클러스터링으로 각 스코어 계산
    : 정확한 분석을 위함
    : 열 추가 + 2
4. 오류파일 입력시 오동작 방지를 위한 코드 개선(완)
  - 반복문, 조건문에서 len max 재확인
    : 회사 PC 코드와 비교

VA8 -> VA9 변경 검토 ('20.08/03 ~ 08/10)
1. 게인자동 삽입기 시퀀스와 동일하게 수정 (완)

2. 엑셀 시트 열 사이즈 조정 (ing)
  - (필요 없는데이터는 최소화)
  
3. 원노트 업로드 및 깃허브 업로드 (ing)

"""


def group_count_seq(count):
    result = 0
    if count == 0:
        result = 3
    elif count == 1:
        result = 13
    elif count == 2:
        result = 10
    elif count == 3:
        result = 2
    elif count == 4:
        result = 7
    elif count == 5:
        result = 4
    elif count == 6:
        result = 12
    elif count == 7:
        result = 6
    elif count == 8:
        result = 0
    elif count == 9:
        result = 8
    elif count == 10:
        result = 9
    elif count == 11:
        result = 1
    elif count == 12:
        result = 5
    elif count == 13:
        result = 11
    return result


def pos_raw_cvt(raw):
    result = "-"
    if raw == 1:
        result = "P"
    elif raw == 3:
        result = "R"
    elif raw == 5:
        result = "N"
    elif raw == 7:
        result = "D"
    return result


angle_offset = 0  # 27.6 + 9.83


def SettingSubPlot32_135(list_time, list_angle, file_name):
    # process ratio : 4.6%, plot ratio : 95.4%

    angle_direction = True  # upper (default)
    overshoot = 0
    time_angle_difference = 0
    overshoot_difference = 0
    time_overshoot_difference = 0
    stop2ovs_difference = 0
    angle_start = 0
    angle_stop = 0
    angle_difference = 0
    count_undershoot = 0
    accuracy = 0

    # 시작, 끝 포인트의 각도에 차이가 1도 이상 있으면
    if abs(float(list_angle[0]) - float(list_angle[-1])) > 1.0:
        if list_angle[0] > list_angle[-1]:
            angle_direction = False  # lower

        # 시작 포인트 탐색 (처음 데이터 기준으로 0.1도 차이)
        for point in range(0, len(list_angle)):
            if abs(float(list_angle[point]) - float(list_angle[0])) > 0.1:
                angle_start = point
                break

        # 끝 포인트 탐색 (끝 데이터 기준으로 0.1도 차이)
        for point in range(len(list_angle) - 1, 0, -1):
            if abs(float(list_angle[point]) - float(list_angle[-1])) > 0.1:
                angle_stop = point
                break

        # 시작, 끝 포인트로부터 각도 차이 (절대값)
        angle_difference = "%5.2f" % (abs(list_angle[angle_stop] - list_angle[angle_start]))

        # 시작, 끝 포인트로부터 시간 차이
        time_angle_difference = "%5.2f" % (list_time[angle_stop] - list_time[angle_start])

        # 상승, 하강에 따른 오버슈트 탐색 (min/max)
        if angle_direction:
            overshoot = list_angle.index(max(list_angle))
        else:
            overshoot = list_angle.index(min(list_angle))

        # 시작부터 오버슈트까지 각도 차이 (절대값)
        overshoot_difference = "%5.2f" % (abs((list_angle[angle_start] - list_angle[overshoot])))

        # 시작부터 오버슈트까지 시간 차이
        time_overshoot_difference = "%5.2f" % (list_time[overshoot] - list_time[angle_start])

        # 끝, 오버슈트와의 각도 차이
        stop2ovs_difference = "%5.2f" % abs(float(angle_difference) - float(overshoot_difference))

        if float(stop2ovs_difference) >= float(2.0):
            gap_set_validity(False, "stop2ovs over 2.0 deg")
        elif float(stop2ovs_difference) >= float(1.5):
            gap_set_validity(False, "stop2ovs over 1.5 deg")
        elif float(stop2ovs_difference) >= float(1.2):
            gap_set_validity(False, "stop2ovs over 1.2 deg")

    if MAKE_A_PLOT_FLAG:
        plt.subplot(3, 2, (1, 5))
        title = " [ " + file_name[1][0] + " 2 " + file_name[1][1] + " ] " + " [ " + file_name[2] + " V ] " + \
                " [ " + file_name[3] + "_" + file_name[4] + "_" + file_name[5] + " ] " + \
                " [ AW ( " + file_name[6] + " ) N ( " + file_name[7] + " ) ] "
        plt.title(title)
        plt.grid(True)
        plt.plot(list_time, list_angle, 'red', linewidth=1.5, alpha=1)  # data
        plt.legend(['ANGLE'], loc='upper right')
        xticks_key = ["0ms", "200ms", "400ms", "600ms", "800ms"]
        xticks_value = [0, 200, 400, 600, 800]
        plt.xlim(0, list_time[-1])

        # N-D : 9.83
        # N-R : 9.83
        # N-P : 27.6

        def_angle_p = 0
        def_angle_r = -17.77
        def_angle_n = -27.60
        def_angle_d = -37.43

        yticks_key = ["P", "-10", "R", "-20", "N", "-30", "D", "-40"]
        yticks_value = [def_angle_p, -10, def_angle_r, -20, def_angle_n, -30, def_angle_d, -40]

        # end point 정확도
        if file_name[1][1] == 'P':
            accuracy = abs(list_angle[angle_stop] - def_angle_p)
        elif file_name[1][1] == 'R':
            accuracy = abs(list_angle[angle_stop] - def_angle_r)
        elif file_name[1][1] == 'N':
            accuracy = abs(list_angle[angle_stop] - def_angle_n)
        elif file_name[1][1] == 'D':
            accuracy = abs(list_angle[angle_stop] - def_angle_d)

        if accuracy > 1.5:
            acc_set_validity(False, "encoder over 1.5 deg")
        elif accuracy > 1.0:
            acc_set_validity(False, "encoder over 1.0 deg")
        elif accuracy > 0.5:
            acc_set_validity(False, "encoder over 0.5 deg")

        # PRND 가이드라인
        plt.annotate("", xy=(list_time[0], def_angle_p), xytext=(list_time[-1], def_angle_p),
                     arrowprops={'arrowstyle': '-', 'color': 'yellow', 'linewidth': 1.5}, alpha=0.5)
        plt.annotate("", xy=(list_time[0], def_angle_r), xytext=(list_time[-1], def_angle_r),
                     arrowprops={'arrowstyle': '-', 'color': 'yellow', 'linewidth': 1.5}, alpha=0.5)
        plt.annotate("", xy=(list_time[0], def_angle_n), xytext=(list_time[-1], def_angle_n),
                     arrowprops={'arrowstyle': '-', 'color': 'yellow', 'linewidth': 1.5}, alpha=0.5)
        plt.annotate("", xy=(list_time[0], def_angle_d), xytext=(list_time[-1], def_angle_d),
                     arrowprops={'arrowstyle': '-', 'color': 'yellow', 'linewidth': 1.5}, alpha=0.5)

        plt.ylim(-50, 10)

        # 시작, 끝 포인트의 각도에 차이가 1도 이상 있으면
        if abs(float(list_angle[0]) - float(list_angle[-1])) > 1.0:

            #
            #
            #
            # VA8 Under Shoot Check Algorithm
            time_undershoot_start = min(angle_stop, overshoot)
            time_undershoot_end = max(angle_stop, overshoot)
            angle_undershoot_low = min(list_angle[angle_stop], list_angle[overshoot]) - 0.5
            angle_undershoot_high = max(list_angle[angle_stop], list_angle[overshoot]) + 0.5

            # 끝까지 탐색 (언더슈트 밴드 외 구간 탐색)
            continuity_flag = 0
            for undershoot in range(time_undershoot_start, len(list_time) - 1):
                if list_time[undershoot] >= 800.0:
                    break
                if undershoot >= time_undershoot_end:
                    break
                if (list_angle[undershoot] < angle_undershoot_low) or (list_angle[undershoot] > angle_undershoot_high):
                    if continuity_flag > 10:
                        plt.plot(list_time[undershoot], list_angle[undershoot], 'g^')
                        count_undershoot = count_undershoot + 1
                    continuity_flag = continuity_flag + 1
                else:
                    # plt.plot(list_time[undershoot], list_angle[undershoot], 'r^')
                    continuity_flag = 0

            if count_undershoot >= 400:  # 40ms
                rng_set_validity(False, "UnderShoot 400ms")
            elif count_undershoot >= 200:  # 20ms
                rng_set_validity(False, "UnderShoot 200ms")
            elif count_undershoot >= 100:  # 10ms
                rng_set_validity(False, "UnderShoot 100ms")
            #
            #
            #

            # 파란 화살표 : 끝위치에서 끝-오버슈트까지 각도
            plt.annotate("", xy=(list_time[angle_stop], list_angle[overshoot]),
                         xytext=(list_time[angle_stop], list_angle[angle_stop]),
                         arrowprops={'arrowstyle': '<->', 'color': 'blue', 'linewidth': 0.5}, alpha=0.5)

            # 파란 텍스트 : 끝위치에서 끝-오버슈트까지 각도
            plt.annotate(stop2ovs_difference + " deg",
                         xy=(list_time[angle_stop] + 10, (list_angle[overshoot] + list_angle[angle_stop]) / 2),
                         ha='left', va='center', fontsize=11, alpha=1.0, color='blue')

            # 검정 화살표 : 시작위치에서 시작-끝까지 각도
            plt.annotate("", xy=(list_time[angle_start], list_angle[angle_start]),
                         xytext=(list_time[angle_start], list_angle[angle_stop]),
                         alpha=1.0, arrowprops={'arrowstyle': '<->', 'color': 'black', 'linewidth': 0.5})

            # 검정 텍스트 : 시작위치에서 시작-끝까지 각도
            plt.annotate(angle_difference + " deg", xy=(
                list_time[angle_start] - 10, (list_angle[angle_start] + list_angle[angle_stop]) / 2),
                         ha='right', va='center', fontsize=11, alpha=1.0, color='black')

            # 빨강 화살표 : 오버슈트 위치에서 시작-오버슈트까지 각도
            plt.annotate("", xy=(list_time[overshoot], list_angle[overshoot]),
                         xytext=(list_time[overshoot], list_angle[angle_start]),
                         arrowprops={'arrowstyle': '<->', 'color': 'red', 'linewidth': 0.5}, alpha=0.5)

            # 빨강 텍스트 : 오버슈트 위치에서 시작-오버슈트까지 각도
            plt.annotate(overshoot_difference + " deg",
                         xy=(list_time[overshoot] + 10, (list_angle[overshoot] + list_angle[angle_start]) / 2),
                         ha='left', va='center', fontsize=11, alpha=1.0, color='red')

            # 상승 그래프의 경우
            if angle_direction:

                # 검정 화살표 : 스타트 포인트 표시
                temp = "Sp ABS\n%.2f deg" % (list_angle[angle_start])
                plt.annotate(temp, size=11, ha="right", va="center", color='black',
                             xy=(list_time[angle_start], list_angle[angle_start]),
                             xytext=(list_time[angle_start] - 20, list_angle[angle_start] - 3),
                             arrowprops=dict(arrowstyle='->', connectionstyle="angle3, angleA=0, angleB=-90",
                                             color='black'))

                # 검정 화살표 : 끝 포인트 표시
                temp = "Ep ABS\n%.2f deg" % (list_angle[angle_stop])
                plt.annotate(temp, size=11, ha="left", va="center", color='black',
                             xy=(list_time[angle_stop], list_angle[angle_stop]),
                             xytext=(list_time[angle_stop] + 20, list_angle[angle_stop] - 3),
                             arrowprops=dict(arrowstyle='->', connectionstyle="angle3, angleA=0, angleB=-90",
                                             color='black'))

                # 검정 양방향 화살표 : 시작-끝 시간 표시
                plt.annotate("", xy=(list_time[angle_start], list_angle[angle_start]),
                             xytext=(list_time[angle_stop], list_angle[angle_start]),
                             arrowprops={'arrowstyle': '<->', 'color': 'black', 'linewidth': 0.5}, alpha=1.0)

                # 검정 양방향 화살표 주석 : 시작-끝 시간 표시
                plt.annotate(time_angle_difference + " ms", xy=(
                    (list_time[angle_stop] + list_time[angle_start]) / 2, list_angle[angle_start] - 3),
                             ha='center', va='bottom', fontsize=11, alpha=1.0, color='black')

                # 빨간 화살표 : 오버슈트 포인트 표시
                temp = "Op ABS\n%.2f deg" % (list_angle[overshoot])
                plt.annotate(temp, size=11, ha="left", va="center", color='red',
                             xy=(list_time[overshoot], list_angle[overshoot]),
                             xytext=(list_time[overshoot] + 20, list_angle[overshoot] + 3),
                             arrowprops=dict(arrowstyle='->', connectionstyle="angle3, angleA=0, angleB=-90",
                                             color='red'))

                # 빨간 양방향 화살표 : 시작-오버슈트 시간 표시
                plt.annotate("", xy=(list_time[angle_start], list_angle[overshoot]),
                             xytext=(list_time[overshoot], list_angle[overshoot]),
                             arrowprops={'arrowstyle': '<->', 'color': 'red', 'linewidth': 0.5}, alpha=0.5)

                # 빨간 양방향 화살표 주석 : 시작-오버슈트 시간 표시
                plt.annotate(time_overshoot_difference + " ms",
                             xy=((list_time[overshoot] + list_time[angle_start]) / 2, list_angle[overshoot] + 3),
                             ha='center', va='top', fontsize=11, alpha=1.0, color='red')

            # 하강 그래프의 경우
            else:

                # 검정 화살표 : 스타트 포인트 표시
                temp = "Sp ABS\n%.2f deg" % (list_angle[angle_start])
                plt.annotate(temp, size=11, ha="right", va="center", color='black',
                             xy=(list_time[angle_start], list_angle[angle_start]),
                             xytext=(list_time[angle_start] - 20, list_angle[angle_start] + 3),
                             arrowprops=dict(arrowstyle='->', connectionstyle="angle3, angleA=0, angleB=-90",
                                             color='black'))
                # 검정 화살표 : 끝 포인트 표시
                temp = "Ep ABS\n%.2f deg" % (list_angle[angle_stop])
                plt.annotate(temp, size=11, ha="left", va="center", color='black',
                             xy=(list_time[angle_stop], list_angle[angle_stop]),
                             xytext=(list_time[angle_stop] + 20, list_angle[angle_stop] + 3),
                             arrowprops=dict(arrowstyle='->', connectionstyle="angle3, angleA=0, angleB=-90",
                                             color='black'))

                # 검정 양방향 화살표 : 시작-끝 시간 표시
                plt.annotate("", xy=(list_time[angle_start], list_angle[angle_start]),
                             xytext=(list_time[angle_stop], list_angle[angle_start]),
                             arrowprops={'arrowstyle': '<->', 'color': 'black', 'linewidth': 0.5}, alpha=1.0)

                # 검정 양방향 화살표 주석 : 시작-끝 시간 표시
                plt.annotate(time_angle_difference + " ms", xy=(
                    (list_time[angle_stop] + list_time[angle_start]) / 2, list_angle[angle_start] + 3),
                             ha='center', va='top', fontsize=11, alpha=1.0, color='black')

                # 빨간 화살표 : 오버슈트 포인트 표시
                temp = "Op ABS\n%.2f deg" % (list_angle[overshoot])
                plt.annotate(temp, size=11, ha="left", va="center", color='red',
                             xy=(list_time[overshoot], list_angle[overshoot]),
                             xytext=(list_time[overshoot] + 20, list_angle[overshoot] - 3),
                             arrowprops=dict(arrowstyle='->', connectionstyle="angle3, angleA=0, angleB=90",
                                             color='red'))

                # 빨간 양방향 화살표 : 시작-오버슈트 시간 표시
                plt.annotate("", xy=(list_time[angle_start], list_angle[overshoot]),
                             xytext=(list_time[overshoot], list_angle[overshoot]),
                             arrowprops={'arrowstyle': '<->', 'color': 'red', 'linewidth': 0.5}, alpha=0.5)

                # 빨간 양방향 화살표 주석 : 시작-오버슈트 시간 표시
                plt.annotate(time_overshoot_difference + " ms",
                             xy=((list_time[overshoot] + list_time[angle_start]) / 2, list_angle[overshoot] - 3),
                             ha='center', va='bottom', fontsize=11, alpha=1.0, color='red')

            # 오버슈트 포인트 그리드
            xticks_key.append(""), xticks_value.append(list_time[overshoot])
            yticks_key.append(""), yticks_value.append(list_angle[overshoot])

            # 시작 포인트 그리드
            xticks_key.append(""), xticks_value.append(list_time[angle_start])
            yticks_key.append(""), yticks_value.append(list_angle[angle_start])

            # 끝 포인트 그리드
            xticks_key.append(""), xticks_value.append(list_time[angle_stop])
            yticks_key.append(""), yticks_value.append(list_angle[angle_stop])

        plt.xticks(xticks_value, xticks_key)
        plt.yticks(yticks_value, yticks_key)

    return [time_angle_difference, overshoot_difference, time_overshoot_difference, stop2ovs_difference,
            count_undershoot, accuracy]


def SettingSubPlot32_2(list_time, list_target, list_actuator):
    spec = 500
	
    color_plot = "black"
    color_xlsx = "FF000000"
    diff_time = 0

    # POS TAR 처음과 끝이 같은지 비교
    target_position_index = 0
    if int(list_target[0]) != int(list_target[-1]):
        for idx in range(0, len(list_target)):
            if int(list_target[0]) != int(list_target[idx]):
                target_position_index = idx
                break

    # ACT TAR 처음과 끝이 같은지 비교
    actuator_position_index = len(list_actuator) - 1
    if int(list_actuator[0]) != int(list_actuator[-1]):
        for idx in range(len(list_actuator) - 1, 0, -1):
            if int(list_actuator[-1]) != int(list_actuator[idx]):
                actuator_position_index = idx + 1
                break

    # 유효한 데이터의 경우 시간 계산
    if (int(list_target[0]) != int(list_target[-1])) and (int(list_actuator[0]) != int(list_actuator[-1])):
        chamber_latency = 20.0  # $%$ 챔버 시간 삭제
        diff_time = "%.2f" % (list_time[actuator_position_index] - list_time[target_position_index] - chamber_latency)
        src = pos_raw_cvt(int(list_target[0]))
        dst = pos_raw_cvt(int(list_actuator[-1]))

        result = tar2act_time_check(src, dst, diff_time)

        spec = result[0]
        color_plot = result[1]
        color_xlsx = result[2]

    if MAKE_A_PLOT_FLAG:
        plt.subplot(3, 2, 2)
        title = "Tar : [ " + pos_raw_cvt(int(list_target[0])) + " 2 " + pos_raw_cvt(
            int(list_target[-1])) + " ]          " + \
                "Act : [ " + pos_raw_cvt(int(list_actuator[0])) + " 2 " + pos_raw_cvt(int(list_actuator[-1])) + " ]"
        plt.title(title)
        plt.grid(True)
        plt.plot(list_time, list_target, 'green', list_time, list_actuator, 'orange', linewidth=1.5, alpha=1)
        plt.legend(['Target Position', 'Actuator Position'], loc='upper right')
        xticks_key = ["0ms", "200ms", "400ms", "600ms", "800ms"]
        xticks_value = [0, 200, 400, 600, 800]
        plt.xlim(0, list_time[-1])
        yticks_key = ["P", "R", "N", "D"]
        yticks_value = [1, 3, 5, 7]
        plt.ylim(-2, 9)

        # 빨간 체크 : TARGET 포인트 표시
        if int(list_target[0]) != int(list_target[-1]):
            plt.annotate("", xy=(int(list_time[target_position_index]), 7.1),
                         xytext=(int(list_time[target_position_index]), 7.3),
                         va="center", ha='center',
                         arrowprops=dict(arrowstyle='->', connectionstyle="arc3", color="red"))

        # 빨간 체크 : ACT 포인트 표시
        if int(list_actuator[0]) != int(list_actuator[-1]):
            plt.annotate("", xy=(int(list_time[actuator_position_index]), 7.1),
                         xytext=(int(list_time[actuator_position_index]), 7.3),
                         va="center", ha='center',
                         arrowprops=dict(arrowstyle='->', connectionstyle="arc3", color="red"))

        # 유효한 데이터의 경우 추가 표시
        if (int(list_target[0]) != int(list_target[-1])) and (int(list_actuator[0]) != int(list_actuator[-1])):
            # 양방향 컬러 화살표
            plt.annotate("", xy=(list_time[target_position_index], 0),  # list_actuator[ActPosIdx_end]),
                         xytext=(list_time[actuator_position_index], 0),  # list_actuator[ActPosIdx_end]),
                         alpha=1.0, arrowprops={'arrowstyle': '<->', 'color': str(color_plot), 'linewidth': 1})

            # 양방향 컬러 화살표 주석
            plt.annotate(diff_time + " ms" + " (spec : " + str(spec) + " ms)",
                         xy=((list_time[target_position_index] + list_time[actuator_position_index]) / 2, 0.5),
                         ha='center', va='center', fontsize=11, alpha=1, color=color_plot)

            # Grid 추가
            xticks_key.append(""), xticks_value.append(list_time[actuator_position_index])
            xticks_key.append(""), xticks_value.append(list_time[target_position_index])

        plt.xticks(xticks_value, xticks_key)
        plt.yticks(yticks_value, yticks_key)

    return [list_time[target_position_index], list_time[actuator_position_index], color_xlsx]


def SettingSubPlot32_4(list_time, list_volt):  # Voltage
    global score_result
    average = 0
    control_start = 0
    control_end = 0

    for avr in range(0, 100):
        average += list_volt[avr] / 100

    for sp in range(0, len(list_volt)):
        if abs(list_volt[sp] - average) >= 0.3:
            control_start = sp
            break

    for ep in range(len(list_volt) - 1, 0, -1):
        if abs(list_volt[ep] - average) >= 0.3:
            control_end = ep
            break

    control_difference = list_time[control_end] - list_time[control_start]
    if control_difference > 500.0:
        score_result += 2

    idx_max = list_volt.index(max(list_volt))
    idx_min = list_volt.index(min(list_volt))
    max_min_diff = "%5.2f" % (abs(list_volt[idx_max] - list_volt[idx_min]))

    if MAKE_A_PLOT_FLAG:
        plt.subplot(3, 2, 4)
        plt.grid(True)
        plt.plot(list_time, list_volt, 'magenta', linewidth=0.5, alpha=1.0)  # data
        plt.legend(['Voltage'], loc='upper right')
        xticks_key = ["0ms", "200ms", "400ms", "600ms", "800ms"]
        xticks_value = [0, 200, 400, 600, 800]
        # Grid 추가
        xticks_key.append(""), xticks_value.append(list_time[control_start])
        xticks_key.append(""), xticks_value.append(list_time[control_end])
        plt.xticks(xticks_value, xticks_key)
        plt.xlim(0, list_time[-1])
        yticks_key = ["9", "12", "16"]
        yticks_value = [9, 12, 16]
        plt.yticks(yticks_value, yticks_key)
        plt.ylim(5, 20)

        # 빨간 체크 : control start point 포인트 표시
        plt.annotate("", xy=(int(list_time[control_start]), 17.1),
                     xytext=(int(list_time[control_start]), 17.3),
                     va="center", ha='center',
                     arrowprops=dict(arrowstyle='->', connectionstyle="arc3", color="red"))

        # 빨간 체크 : control end point 포인트 표시
        plt.annotate("", xy=(int(list_time[control_end]), 17.1),
                     xytext=(int(list_time[control_end]), 17.3),
                     va="center", ha='center',
                     arrowprops=dict(arrowstyle='->', connectionstyle="arc3", color="red"))

        # 단방향 자주 화살표 : 최대 전압
        temp = "%.2f V" % (list_volt[idx_max])
        plt.annotate(temp, size=11, ha="left", va="center", color='magenta',
                     xy=(list_time[idx_max], list_volt[idx_max]),
                     xytext=(list_time[idx_max] + 20, 18),  # list_volt[idx_max] + 3),
                     arrowprops=dict(arrowstyle='->', connectionstyle="angle3, angleA=0, angleB=-90",
                                     color='magenta'))

        # 단방향 자주 화살표 : 최소 전압
        temp = "%.2f V" % (list_volt[idx_min])
        plt.annotate(temp, size=11, ha="left", va="center", color='magenta',
                     xy=(list_time[idx_min], list_volt[idx_min]),
                     xytext=(list_time[idx_min] + 20, 7),  # list_volt[idx_min] - 3),
                     arrowprops=dict(arrowstyle='->', connectionstyle="angle3, angleA=0, angleB=-90",
                                     color='magenta'))

        # 양방향 검정 화살표 : P2P 전압
        plt.annotate("", xy=(list_time[50], list_volt[idx_min]),
                     xytext=(list_time[50], list_volt[idx_max]),
                     arrowprops={'arrowstyle': '<->', 'color': 'black', 'linewidth': 0.5}, alpha=0.5)

        # 양방향 검정 화살표 주석: P2P 전압
        plt.annotate(max_min_diff + " V",
                     xy=(list_time[50] + 10, (list_volt[idx_min] + list_volt[idx_max]) / 2),
                     ha='left', va='center', fontsize=11, alpha=1.0, color='black')

    return [list_volt[idx_max], list_volt[idx_min], max_min_diff, control_difference]


def SettingSubPlot32_6(list_time, list_curr):  # Current

    idx_max = list_curr.index(max(list_curr))
    idx_min = list_curr.index(min(list_curr))
    max_min_diff = "%5.2f" % (abs(list_curr[idx_max] - list_curr[idx_min]))

    if MAKE_A_PLOT_FLAG:
        plt.subplot(3, 2, 6)
        plt.grid(True)
        plt.plot(list_time, list_curr, 'blue', linewidth=0.5, alpha=1.0)  # data
        plt.legend(['Current'], loc='upper right')
        xticks_key = ["0ms", "200ms", "400ms", "600ms", "800ms"]
        xticks_value = [0, 200, 400, 600, 800]
        plt.xticks(xticks_value, xticks_key)
        plt.xlim(0, list_time[-1])
        yticks_key = ["-10", "-5", "0", "5", "10"]
        yticks_value = [-10, -5, 0, 5, 10]
        plt.yticks(yticks_value, yticks_key)
        plt.ylim(-15, 15)

        # 단방향 파랑 화살표 : 최대 전류
        temp = "%.2f A" % (list_curr[idx_max])
        plt.annotate(temp, size=11, ha="left", va="center", color='blue',
                     xy=(list_time[idx_max], list_curr[idx_max]),
                     xytext=(list_time[idx_max] + 20, 13),  # list_volt[idx_max] + 3),
                     arrowprops=dict(arrowstyle='->', connectionstyle="angle3, angleA=0, angleB=-90",
                                     color='blue'))

        # 단방향 파랑 화살표 : 최소 전류
        temp = "%.2f A" % (list_curr[idx_min])
        plt.annotate(temp, size=11, ha="left", va="center", color='blue',
                     xy=(list_time[idx_min], list_curr[idx_min]),
                     xytext=(list_time[idx_min] + 20, -13),  # list_volt[idx_min] - 3),
                     arrowprops=dict(arrowstyle='->', connectionstyle="angle3, angleA=0, angleB=-90",
                                     color='blue'))

        # 양방향 검정 화살표 : P2P 전압
        plt.annotate("", xy=(list_time[50], list_curr[idx_min]),
                     xytext=(list_time[50], list_curr[idx_max]),
                     arrowprops={'arrowstyle': '<->', 'color': 'black', 'linewidth': 0.5}, alpha=0.5)

        # 양방향 검정 화살표 주석 : P2P 전압
        plt.annotate(max_min_diff + " A",
                     xy=(list_time[50] + 10, (list_curr[idx_min] + list_curr[idx_max]) / 2),
                     ha='left', va='center', fontsize=11, alpha=1.0, color='black')

    return [list_curr[idx_max], list_curr[idx_min], max_min_diff]


def SettingSubPlot42_8(list_time, list_curr_state_u, list_curr_state_v, list_curr_state_w):  # Current State UVW
    plt.subplot(4, 2, 8)
    plt.grid(True)
    plt.plot(list_time, list_curr_state_u, 'coral', list_time, list_curr_state_v, 'green', list_time, list_curr_state_w,
             'mediumslateblue', linewidth=1.0, alpha=1.0)  # data
    plt.legend(['Curr_U', 'Curr_V', 'Curr_W'], loc='upper right')
    xticks_key = ["0ms", "200ms", "400ms", "600ms", "800ms"]
    xticks_value = [0, 200, 400, 600, 800]
    plt.xticks(xticks_value, xticks_key)
    plt.xlim(0, list_time[-1])

    yticks_key = ["5/div", "W", "", "", "V", "", "", "U", ""]  # , "W", "5", "-5", "V", "5", "-5", "U", "5"]
    yticks_value = [-20, -15, -10, -5, 0, 5, 10, 15, 20]
    plt.yticks(yticks_value, yticks_key)
    plt.ylim(-25, 25)

    return [1, 2, 3]


def median_flt(data):
    median = []
    for point in range(0, 800, 40):  # 0 ~ 800 : 20 sample
        median.append(data[point])
    data[0] = statistics.median(median)


WB = Workbook()  # Create execl file
WS = WB.active
WS.title = "ChangAn"  # Worksheet Title

Description = ["FileName", "group", "seq", "DIR", "VOLT", "P", "I", "D", "AW", "N",  # 0 ~ 9
"MxM SPD", "M N SPD", "MxM CURR",  # 10 ~ 12
"[Score]",  # 13
"Stable Time", "OvS Angle", "OvS Time",  # 14 ~ 16
"Gap (Stable-OvS)", "Gap Validity",  # 17 ~ 18
"Tar2Act Time", "Tar2Act Validity",  # 19 ~ 20
"Out of RNG", "RNG Validity",  # 21 ~ 22
"encoder accuracy", "acc Validity",  # 23 ~ 24
"Control Time",  # 25
"Max Voltage", "Min Voltage", "P2P Voltage",  # 26 ~ 28
"Max Ampere", "Min Ampere", "P2P Ampere", "Graph Link"]  # 29 ~ 32

for col in range(0, len(Description)):
    WS.cell(0 + 1, col + 1).value = Description[col]

    score_color = PatternFill(start_color="FF888888", fill_type="solid")  # gray
    WS.cell(0 + 1, 13 + 1).fill = score_color
    WS.cell(0 + 1, 17 + 1).fill = score_color
    WS.cell(0 + 1, 18 + 1).fill = score_color
    WS.cell(0 + 1, 19 + 1).fill = score_color
    WS.cell(0 + 1, 20 + 1).fill = score_color
    WS.cell(0 + 1, 21 + 1).fill = score_color
    WS.cell(0 + 1, 22 + 1).fill = score_color
    WS.cell(0 + 1, 23 + 1).fill = score_color
    WS.cell(0 + 1, 24 + 1).fill = score_color
    WS.cell(0 + 1, 25 + 1).fill = score_color

direction_list = ["PR", "PN", "PD", "RP", "RN", "RD", "NP", "NR", "ND", "DP", "DR", "DN"]

"""
프로그램 시작
"""

valid_file_name_list = []
valid_file_addr_list = []

FolderLevel1Name = os.path.basename(os.getcwd())
FolderLevel1Path = os.getcwd()
print("(Lv1) %s" % FolderLevel1Name, end="\n")

for FolderLevel2Name in os.listdir(os.getcwd()):
    FolderLevel2Path = os.path.join(os.getcwd(), FolderLevel2Name)
if os.path.isdir(FolderLevel2Path):
    print("(Lv2)   └ %s" % FolderLevel2Name, end="\n")

for FolderLevel3Name in os.listdir(FolderLevel2Path):
    FolderLevel3Path = os.path.join(FolderLevel2Path, FolderLevel3Name)
if os.path.isdir(FolderLevel3Path):
    print("(Lv3)     └ %s" % FolderLevel3Name, end="\n")

for FileName in os.listdir(FolderLevel3Path):
    FilePath = os.path.join(FolderLevel3Path, FileName)
    if os.path.isfile(FilePath):
        if re.findall(".csv", FileName):  # 파일 익스텐션 체크
            if len(FileName.split('_')) == 11:  # 토큰 개수 체크
                print("(Lv4)       └ %-56s " % FileName, end="\n")  # 파일 이름 출력
                valid_file_name_list.append(FileName.split('.')[0])  # 파일 익스텐션 삭제
                valid_file_addr_list.append(FilePath)
    else:
        print("(Lv4)       └ %-56s : not matching number of '_' token " % FileName, end="\n")

"""
결과 폴더 확인 및 생성
"""

if not os.path.isdir("Good Graph Result"):
    os.mkdir("Good Graph Result")

for folder in range(0, len(direction_list)):
    if not os.path.isdir("Good Graph Result\\" + direction_list[folder]):
        os.mkdir("Good Graph Result\\" + direction_list[folder])

if not os.path.isdir("Bad Graph Result"):
    os.mkdir("Bad Graph Result")

for folder in range(0, len(direction_list)):
    if not os.path.isdir("Bad Graph Result\\" + direction_list[folder]):
        os.mkdir("Bad Graph Result\\" + direction_list[folder])

CSVCounter = len(valid_file_name_list)
print("Number of CSV files : %4d" % CSVCounter, end="\n")

program_start = time.time()

for index_file in range(0, len(valid_file_name_list)):

    threading_variable_file = index_file + 1
    print('\r processed CSV files : %4s ' % str(threading_variable_file), end='')

gap_validity = True
gap_validity_text = ""
t2a_validity = True
t2a_validity_text = ""
rng_validity = True
rng_validity_text = ""
acc_validity = True
acc_validity_text = ""

FileName = valid_file_name_list[index_file]
LoggingDate = FileName.split('_')[0][0:8]
LoggingTime = FileName.split('_')[0][8:14]
MotorDirection = FileName.split('_')[1]

data_time = []
data_current = []
data_voltage = []
data_angle = []
data_target = []
data_actuator = []
data_current_state_u = []
data_current_state_v = []
data_current_state_w = []

log_time_limit = 800.0  # 그래프 리미트
log_line_count = 0

with open(valid_file_addr_list[index_file], 'r') as file:
    log = csv.reader(file)
    column = next(log)

    index_time = column.index("Time[ms]")
    index_current = column.index("I_DC")
    index_voltage = column.index("V_DC")
    index_angle = column.index("Angle")
    index_target = column.index("POS TAR")
    index_actuator = column.index("POS ACT")

    try:
        index_current_state_u = column.index("I_U")
    except ValueError:
        index_current_state_u = -1

    try:
        index_current_state_v = column.index("I_V")
    except ValueError:
        index_current_state_v = -1

    try:
        index_current_state_w = column.index("I_W")
    except ValueError:
        index_current_state_w = -1

    for raw_data in log:

        log_line_count = log_line_count + 1

        data_time.append(float(raw_data[index_time]))
        data_current.append(float(raw_data[index_current]))
        data_voltage.append(float(raw_data[index_voltage]))
        data_angle.append(float(raw_data[index_angle]) - float(angle_offset))
        data_actuator.append(float(raw_data[index_actuator]))

        target_conversion = int((float(raw_data[index_target]) * 10) / 10)
        if target_conversion == 8:
            data_target.append(float(1.0))
        elif target_conversion == 7:
            data_target.append(float(3.0))
        elif target_conversion == 6:
            data_target.append(float(5.0))
        elif target_conversion == 5:
            data_target.append(float(7.0))

        if data_current_state_u != -1:
            data_current_state_u.append(float(raw_data[index_current_state_u]) + 15.0)  # offset
        else:
            data_current_state_u.append(float(0.0))

        if data_current_state_v != -1:
            data_current_state_v.append(float(raw_data[index_current_state_v]))
        else:
            data_current_state_u.append(float(0.0))

        if data_current_state_w != -1:
            data_current_state_w.append(float(raw_data[index_current_state_w]) - 15.0)  # offset
        else:
            data_current_state_u.append(float(0.0))

        if float(raw_data[index_time]) > log_time_limit:
            break

# MEDIAN FILTER _ Find Start Point
median_flt(data_current)
median_flt(data_voltage)
median_flt(data_angle)
median_flt(data_current_state_u)
median_flt(data_current_state_v)
median_flt(data_current_state_w)

current_over_limit = 40.0
for previous in range(1, log_line_count):
    if data_voltage[previous] >= current_over_limit:
        data_current[previous] = data_current[previous - 1]
        data_voltage[previous] = data_voltage[previous - 1]
        data_angle[previous] = data_angle[previous - 1]
        data_current_state_u[previous] = data_current_state_u[previous - 1]
        data_current_state_v[previous] = data_current_state_v[previous - 1]
        data_current_state_w[previous] = data_current_state_w[previous - 1]

if MAKE_A_PLOT_FLAG:
    plt.figure(figsize=(19, 9.5))

    FileNameList = FileName.split('.')[0].split('_')
    Angle = SettingSubPlot32_135(data_time, data_angle, FileNameList)
    Pos = SettingSubPlot32_2(data_time, data_target, data_actuator)
    Voltage = SettingSubPlot32_4(data_time, data_voltage)
    Current = SettingSubPlot32_6(data_time, data_current)

    if MAKE_A_PLOT_FLAG:
        plt.tight_layout()

    graph_file_link = ""
    for folder in range(0, len(direction_list)):
        if (FileNameList[1][0] == direction_list[folder][0]) and \
                (FileNameList[1][1] == direction_list[folder][1]):
            if gap_validity and t2a_validity and rng_validity:
                if MAKE_A_PLOT_FLAG:
                    plt.savefig("Good Graph Result\\" + direction_list[folder] + "\\" + FileName)
                    plt.close()
                graph_file_link = FolderLevel1Path + "\\Good Graph Result\\" + direction_list[
                    folder] + "\\" + FileName + ".png"

            else:
                if MAKE_A_PLOT_FLAG:
                    plt.savefig("Bad Graph Result\\" + direction_list[folder] + "\\" + FileName)
                    plt.close()
                graph_file_link = FolderLevel1Path + "\\Bad Graph Result\\" + direction_list[
                    folder] + "\\" + FileName + ".png"
            break

    WS.cell(threading_variable_file + 1, 0 + 1).value = FileNameList[0]

    WS.cell(threading_variable_file + 1, 1 + 1).value = '{0:03d}'.format(int((threading_variable_file - 1) / 14))
    file_group_count = (threading_variable_file - 1) % 14  # sort
    WS.cell(threading_variable_file + 1, 2 + 1).value = group_count_seq(file_group_count)
    WS.cell(threading_variable_file + 1, 3 + 1).value = FileNameList[1]
    WS.cell(threading_variable_file + 1, 4 + 1).value = FileNameList[2]
    WS.cell(threading_variable_file + 1, 5 + 1).value = FileNameList[3]
    WS.cell(threading_variable_file + 1, 6 + 1).value = FileNameList[4]
    WS.cell(threading_variable_file + 1, 7 + 1).value = FileNameList[5]
    WS.cell(threading_variable_file + 1, 8 + 1).value = FileNameList[6]
    WS.cell(threading_variable_file + 1, 9 + 1).value = FileNameList[7]
    WS.cell(threading_variable_file + 1, 10 + 1).value = FileNameList[8]
    WS.cell(threading_variable_file + 1, 11 + 1).value = FileNameList[9]
    WS.cell(threading_variable_file + 1, 12 + 1).value = FileNameList[10]
    WS.cell(threading_variable_file + 1, 13 + 1).value = "---"
    WS.cell(threading_variable_file + 1, 14 + 1).value = float(Angle[0])
    WS.cell(threading_variable_file + 1, 15 + 1).value = float(Angle[1])
    WS.cell(threading_variable_file + 1, 16 + 1).value = float(Angle[2])

    Gap = float(Angle[3])
    if Gap >= float(2.0):
        score_result += 3
        color_gap = PatternFill(start_color="FFFF2211", fill_type="solid")  # red
    elif Gap >= float(1.5):
        score_result += 2
        color_gap = PatternFill(start_color="FFFF9911", fill_type="solid")  # orange
    elif Gap >= float(1.2):
        score_result += 1
        color_gap = PatternFill(start_color="FFAA9148", fill_type="solid")  # gold
    else:
        color_gap = PatternFill(start_color="FF44FF33", fill_type="solid")  # forest green
    WS.cell(threading_variable_file + 1, 17 + 1).fill = color_gap
    WS.cell(threading_variable_file + 1, 17 + 1).value = Gap
    WS.cell(threading_variable_file + 1, 18 + 1).value = gap_validity_text

    Tar2Act = "%.2f" % (Pos[1] - Pos[0])
    color_tar2act = PatternFill(start_color=Pos[2], fill_type="solid")
    WS.cell(threading_variable_file + 1, 19 + 1).fill = color_tar2act
    WS.cell(threading_variable_file + 1, 19 + 1).value = float(Tar2Act)
    WS.cell(threading_variable_file + 1, 20 + 1).value = t2a_validity_text

    if Angle[4] >= 400:  # 40ms
        score_result += 10
        color_undershoot = PatternFill(start_color="FFFF2211", fill_type="solid")
    elif Angle[4] >= 200:  # 20ms
        score_result += 5
        color_undershoot = PatternFill(start_color="FFFF9911", fill_type="solid")
    elif Angle[4] >= 100:  # 10ms
        score_result += 3
        color_undershoot = PatternFill(start_color="FFAA9148", fill_type="solid")
    else:
        color_undershoot = PatternFill(start_color="FF44FF33", fill_type="solid")  # forest green
    WS.cell(threading_variable_file + 1, 21 + 1).fill = color_undershoot
    WS.cell(threading_variable_file + 1, 21 + 1).value = Angle[4]  # Undershoot
    WS.cell(threading_variable_file + 1, 22 + 1).value = rng_validity_text

    if Angle[5] > 1.5:  # encoder over 2 deg
        score_result += 3
        color_accuracy = PatternFill(start_color="FFFF2211", fill_type="solid")
    elif Angle[5] > 1.0:  # encoder over 1 deg
        score_result += 2
        color_accuracy = PatternFill(start_color="FFFF9911", fill_type="solid")
    elif Angle[5] > 0.5:  # encoder over 1 deg
        score_result += 1
        color_accuracy = PatternFill(start_color="FFAA9148", fill_type="solid")
    else:
        color_accuracy = PatternFill(start_color="FF44FF33", fill_type="solid")  # forest green
    WS.cell(threading_variable_file + 1, 23 + 1).fill = color_accuracy
    WS.cell(threading_variable_file + 1, 23 + 1).value = Angle[5]  # accuracy
    WS.cell(threading_variable_file + 1, 24 + 1).value = acc_validity_text  # accuracy

    WS.cell(threading_variable_file + 1, 25 + 1).value = float(Voltage[3])
    WS.cell(threading_variable_file + 1, 26 + 1).value = float(Voltage[0])
    WS.cell(threading_variable_file + 1, 27 + 1).value = float(Voltage[1])
    WS.cell(threading_variable_file + 1, 28 + 1).value = float(Voltage[2])
    WS.cell(threading_variable_file + 1, 29 + 1).value = float(Current[0])
    WS.cell(threading_variable_file + 1, 30 + 1).value = float(Current[1])
    WS.cell(threading_variable_file + 1, 31 + 1).value = float(Current[2])

    WS.cell(threading_variable_file + 1, 32 + 1).value = graph_file_link
    WS.cell(threading_variable_file + 1, 32 + 1).hyperlink = graph_file_link

    if (threading_variable_file - 1) % 14 == 13:
        for rollback in range(0, 14):
            WS.cell(threading_variable_file - rollback + 1, 13 + 1).value = score_result
        score_result = 0

WB.save(FolderLevel1Name + " 20" + datetime.datetime.now().strftime('%m%d %H%M%S') + ".xlsx")

print("\nrunning time : " + str(datetime.timedelta(seconds=time.time() - program_start)).split(".")[0])
