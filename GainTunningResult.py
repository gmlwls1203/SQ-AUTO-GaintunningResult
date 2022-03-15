import time
start_time = time.time()  # 시작 시간 저장
import csv
import os
import pandas as pd
from openpyxl import Workbook
from matplotlib import pyplot as plt

# Ver0315
# - 플롯에 점수 계산 기준 표시
# - 실행 시간 출력
# - 점수가 음수값이면 0 으로 저장

RT_Score = []
# Response time score calculation
def CalScore_RT(raw_data) :
    raw_data[1] = int(raw_data[1])
    raw_data[2] = int(raw_data[2])
    score_val = 0
    spec_val = 0

    if raw_data[1] == 0 : #PR
        spec_val = 250
    elif raw_data[1] == 1: #PN
        spec_val = 300
    elif raw_data[1] == 2: #PD
        spec_val = 450
    elif raw_data[1] == 3: #RP
        spec_val = 1000
    elif raw_data[1] == 4: #RN
        spec_val = 200
    elif raw_data[1] == 5:  # RD
        spec_val = 350
    elif raw_data[1] == 6:  # NP
        spec_val = 300
    elif raw_data[1] == 7:  # NR
        spec_val = 200
    elif raw_data[1] == 8:  # ND
        spec_val = 200
    elif raw_data[1] == 9:  # DP
        spec_val = 1000
    elif raw_data[1] == 10:  # DR
        spec_val = 350
    elif raw_data[1] == 11:  # DN
        spec_val = 200

    if raw_data[2] > spec_val:
        score_val = 100-abs(raw_data[2] - spec_val)
        if score_val < 0 :
            score_val = 0
    else:
        score_val = 100

    list = [int(raw_data[0]), score_val]
    RT_Score.append(list)


SA_Score = []
# Stop Accuracy score calculation
def CalScore_SA(raw_data) :
    score_val = 0
    raw_data[4] = float(raw_data[4])

    if raw_data[4] == 0 :
        score_val = 100
    else :
        score_val = 100 - abs(raw_data[4]) * 10

    list = [int(raw_data[0]), score_val]
    SA_Score.append(list)

OS_Score = []
# Overshoot score calculation
def CalScore_OS(raw_data) :
    raw_data[6] = float(raw_data[6])

    if raw_data[6] == 0 :
        score_val = 100
    else :
        score_val = 100 - abs(raw_data[6]) * 10

    list = [int(raw_data[0]), score_val]
    OS_Score.append(list)


def FindBestGainSet(sum_score):
    max = 0
    for idx in sum_score :
        val = int(idx[1])
        if max < val :
            max = val
            gainset = int(idx[0])
        else:
            continue
    print("[%d] gainset (score : %d) is the Best Gain Set!" %(gainset,max))

# 엑셀 게인값 불러오기
df = pd.read_csv('./gaintuning.csv',usecols=['GroupNum', 'Type', 'Pos_P', 'Pos_I', 'Pos_D', 'Pos_AntiW', 'Pos_Term', 'Spd_P', 'Spd_I', 'Spd_AntiW', 'Curr_P', 'Curr_I', 'Curr_AntiW'])

# 엑셀 데이터 저장
WB = Workbook()
WS = WB.active
WS.title = "SQ_Auto_Gaintunning"

Description = ["Gain Set", "항목", "PR", "PN", "PD", "RP", "RN", "RD", "NP", "NR", "ND", "DP", "DR", "DN", "총점수", "MAX", "MIN"]
for col in range(0, len(Description)):
    WS.cell(0+1, col+1).value = Description[col]

sum_score_RT = 0
sum_score_SA = 0
sum_score_OS = 0
log_count = 0
sum_score = []
row = 0
# read excel file
with open('./TEST_RESULT.csv', 'r') as file :
    log = csv.reader(file)
    for raw_data in log :
        if log_count >= 1 :
            CalScore_RT(raw_data)
            CalScore_SA(raw_data)
            CalScore_OS(raw_data)

        log_count = log_count + 1

    start = 0
    end = len(RT_Score)
    div = 12

    for idx in range(start, end, div):
        resRT = RT_Score[start:start+div]
        resSA = SA_Score[start:start+div]
        resOS = OS_Score[start:start+div]

        # gain set 별 제어 성능 점수 계산
        if resRT != [] :
            val_list = []
            for i in range(len(resRT)) :
                gainset = resRT[i][0]
                val = resRT[i][1]
                sum_score_RT += val
                val_list.append(val)

            max_RT = max(val_list)
            min_RT = min(val_list)
            print("[%d] Gainset Response Time Score : %d" % (gainset,sum_score_RT))

        if resSA != [] :
            val_list = []
            for i in range(len(resSA)) :
                gainset = resSA[i][0]
                val = resSA[i][1]
                sum_score_SA += val
                val_list.append(val)

            max_SA = max(val_list)
            min_SA = min(val_list)
            print("[%d] Gainset Stop Accuracy Score : %d" % (gainset, sum_score_SA))

        if resOS != []:
            val_list = []
            for i in range(len(resOS)):
                gainset = resOS[i][0]
                val = resOS[i][1]
                sum_score_OS += val
                val_list.append(val)

            max_OS = max(val_list)
            min_OS = min(val_list)
            print("[%d] Gainset Overshoot Score : %d" % (gainset, sum_score_OS))

        # Excel 에 data 저장
        col = 0
        WS.append([resRT[col][0], "응답시간", resRT[col][1], resRT[col + 1][1], resRT[col + 2][1], resRT[col + 3][1],
                   resRT[col + 4][1], resRT[col + 5][1], resRT[col + 6][1], resRT[col + 7][1], resRT[col + 8][1],
                   resRT[col + 9][1], resRT[col + 10][1], resRT[col + 11][1],sum_score_RT, max_RT, min_RT])

        WS.append([resSA[col][0], "제어정밀도", resSA[col][1], resSA[col + 1][1], resSA[col + 2][1], resSA[col + 3][1],
                   resSA[col + 4][1], resSA[col + 5][1], resSA[col + 6][1], resSA[col + 7][1], resSA[col + 8][1],
                   resSA[col + 9][1], resSA[col + 10][1], resSA[col + 11][1],sum_score_SA, max_SA, min_SA])

        WS.append([resOS[col][0], "오버슈트", resOS[col][1], resOS[col + 1][1], resOS[col + 2][1], resOS[col + 3][1],
                   resOS[col + 4][1], resOS[col + 5][1], resOS[col + 6][1], resOS[col + 7][1], resOS[col + 8][1],
                   resOS[col + 9][1], resOS[col + 10][1], resOS[col + 11][1],sum_score_OS, max_OS, min_OS])

        list = [gainset, int(sum_score_OS + sum_score_RT + sum_score_SA)]
        sum_score.append(list)
        sum_score_RT = 0
        sum_score_SA = 0
        sum_score_OS = 0
        start = start + div

    FindBestGainSet(sum_score)

# 결과 데이터 플롯 파일 생성 및 엑셀파일에 추가
if not os.path.isdir("SaveFig") :
    os.mkdir("SaveFig")

imgcnt = 0
imgflag = 0
xlabels = ['PR', 'PN', 'PD', 'RP', 'RN', 'RD', 'NP', 'NR', 'ND', 'DP', 'DR', 'DN']
lastcol = WS.max_column
# 행 데이터 확인
for row in WS.iter_rows(min_row = 2, min_col = 2, max_col = 14):
    pltval = []
    # graph 크기 수정, 해상도 설정
    plt.figure(figsize = (9,6), dpi = 200)
    plt.axhline(linewidth=0.7, y=100, color = 'lightgray', linestyle = '--')
    plt.xticks([0,1,2,3,4,5,6,7,8,9,10,11],xlabels)
    plt.yticks(range(-200,201,10),size = 5)
    plt.ylim([0, 120])

    for cell in row:
        if cell.value == "응답시간" :
            plt.title('Response Time Score')
            plt.axhline(linewidth=0.2, y=90, color='orange', linestyle = '--')
            plt.text(11.6, 90, '<-Response time Over 10ms', fontsize=4, va='center')
            plt.axhline(linewidth=0.2, y=80, color='orange', linestyle = '--')
            plt.text(11.6, 80, '<-Response time Over 20ms', fontsize=4, va='center')
            plt.axhline(linewidth=0.2, y=70, color='orange', linestyle = '--')
            plt.text(11.6, 70, '<-Response time Over 30ms', fontsize=4, va='center')
        elif cell.value == "제어정밀도" :
            plt.title('Stop Accuracy Score')
            plt.axhline(linewidth=0.2, y=90, color='orange', linestyle = '--')
            plt.text(11.6, 90, '<-Stop Accuracy Over 1.0%', fontsize=4, va='center')
            plt.axhline(linewidth=0.2, y=80, color='orange', linestyle = '--')
            plt.text(11.6, 80, '<-Stop Accuracy Over 2.0%', fontsize=4, va='center')
            plt.axhline(linewidth=0.2, y=70, color='orange', linestyle = '--')
            plt.text(11.6, 70, '<-Stop Accuracy Over 3.0%', fontsize=4, va='center')
        elif cell.value == "오버슈트" :
            plt.title('Overshoot Score')
            plt.axhline(linewidth=0.2, y=90, color='orange')
            plt.text(11.6, 90, '<-Overshoot Over 1.0%', fontsize=4, va='center')
            plt.axhline(linewidth=0.2, y=80, color='orange', linestyle = '--')
            plt.text(11.6, 80, '<-Overshoot Over 2.0%', fontsize=4, va='center')
            plt.axhline(linewidth=0.2, y=70, color='orange', linestyle = '--')
            plt.text(11.6, 70, '<-Overshoot Over 3.0%', fontsize=4, va='center')
        else :
            pltval.append(cell.value)

    # 그래프 출력
    plt.plot(pltval, '.')

    # Annotation
    x_pos = 0
    for i in range(len(pltval)) :
        if x_pos == 12:
            x_pos = 0
        plt.text(x_pos, pltval[i]+1, int(pltval[i]), fontsize=5)
        x_pos = x_pos+1

    imgcnt = imgcnt + 1

    # 플롯에 제어 시 적용된 게인 표 추가
    gbnum = WS.cell(row=imgcnt+1, column=1).value
    gb = df.groupby('GroupNum').get_group(gbnum)
    columns = gb.columns
    tuples = [tuple(x) for x in gb.to_numpy()]
    plt.table(cellText = tuples, colLabels=columns, loc="bottom", bbox=[0.0, -1.2, 1.0, 1.0], cellLoc='center')

    # table 공간
    plt.gcf().subplots_adjust(bottom=0.5)
    plt.savefig('SaveFig\\img_'+str(imgcnt)+'.png')
    graph_file_link = 'SaveFig\\img_'+str(imgcnt)+'.png'
    plt.close()

    if imgflag == 0:
        imgflag = 1
        WS.cell(imgflag, lastcol+1).value = "그래프"

    WS.cell(imgcnt+1, lastcol+1).value = graph_file_link
    WS.cell(imgcnt+1, lastcol+1).hyperlink = "./" + graph_file_link

WB.save("SQ GainTunning Test Result.xlsx")

print("time : ", format(time.time()-start_time, ".2f"))